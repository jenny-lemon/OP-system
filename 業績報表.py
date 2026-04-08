import os
import sys
import calendar
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from typing import Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from accounts import ACCOUNTS
from paths import PATH_JENNY


LOGIN_URL = "https://backend.lemonclean.com.tw/login"
PURCHASE_URL = "https://backend.lemonclean.com.tw/purchase"

HEADERS = {"User-Agent": "Mozilla/5.0"}

CITY_ORDER = ["台北", "台中", "桃園", "新竹", "高雄"]
INCOME_ORDER = ["現金收入", "儲值金"]

CATEGORY_ORDER = ["清潔", "儲值金", "冷氣", "洗衣機", "水洗", "收納"]

REGION3_CATEGORY_ORDER = [
    "清潔",
    "冷氣",
    "洗衣機",
    "水洗",
    "收納",
    "儲值金",
    "清潔現金+儲值金",
    "家電現金+儲值金",
    "水洗/收納現金+儲值金",
    "清潔+水洗+收納現金+儲值金",
]


def login(session, email, password):
    res = session.get(LOGIN_URL, headers=HEADERS, allow_redirects=True)
    res.raise_for_status()

    soup = BeautifulSoup(res.text, "html.parser")
    token_input = soup.find("input", {"name": "_token"})
    if not token_input:
        raise RuntimeError("找不到 _token，無法登入")

    payload = {
        "_token": token_input.get("value"),
        "email": email,
        "password": password,
    }

    res = session.post(LOGIN_URL, data=payload, headers=HEADERS, allow_redirects=True)
    res.raise_for_status()

    if "login" in res.url.lower():
        raise RuntimeError(f"{email} 登入失敗")

    print(f"✅ 登入成功：{email}")


def get_ranges():
    today = datetime.today()
    y, m = today.year, today.month

    this_start = f"{y}-{m:02d}-01"
    this_end = f"{y}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"

    if m == 12:
        ny, nm = y + 1, 1
    else:
        ny, nm = y, m + 1

    next_start = f"{ny}-{nm:02d}-01"
    next_end = f"{ny}-{nm:02d}-{calendar.monthrange(ny, nm)[1]:02d}"

    return (this_start, this_end), (next_start, next_end)


def build_url(start, end, status, keyword=""):
    params = {
        "keyword": keyword,
        "name": "",
        "phone": "",
        "orderNo": "",
        "date_s": "",
        "date_e": "",
        "clean_date_s": start,
        "clean_date_e": end,
        "paid_at_s": "",
        "paid_at_e": "",
        "refundDateS": "",
        "refundDateE": "",
        "buy": "",
        "area_id": "",
        "isCharge": "",
        "isRefund": "",
        "p_board": "on",
        "payway": "",
        "purchase_status": str(status),
        "progress_status": "",
        "invoiceStatus": "",
        "otherFee": "",
        "orderBy": "",
    }
    return requests.Request("GET", PURCHASE_URL, params=params).prepare().url


def get_keywords(city):
    if city == "新竹":
        return ["新竹"]
    if city == "高雄":
        return ["高雄", "台南"]
    return [""]


def safe_int(v):
    try:
        s = str(v).replace(",", "").strip()
        if s in ("", "-", "None", "nan"):
            return 0
        return int(float(s))
    except Exception:
        return 0


def normalize_service(name):
    name = str(name or "").strip()
    name = name.replace("螨", "蟎")

    mapping = {
        "VIP": "儲值金",
        "冷氣機清潔": "冷氣清潔",
        "冷氣機清潔服務": "冷氣清潔",
        "洗衣機": "洗衣機清潔",
        "沙發床墊水洗除蟎": "水洗",
        "沙發床墊水洗除螨": "水洗",
        "沙發清洗": "水洗",
        "床墊清洗": "水洗",
        "整理收納": "收納",
    }
    return mapping.get(name, name)


def detect_income_type(first_header):
    first_header = str(first_header or "").strip()
    if first_header in ("VIP", "儲值金"):
        return "儲值金"
    return "現金收入"


def parse_html(html):
    soup = BeautifulSoup(html, "html.parser")
    tables = soup.find_all("table")
    results = []

    for table in tables:
        trs = table.find_all("tr")
        rows = []

        for tr in trs:
            cells = tr.find_all(["th", "td"])
            row = [c.get_text(" ", strip=True) for c in cells]
            if any(str(x).strip() for x in row):
                rows.append(row)

        if not rows:
            continue

        header = [str(x).strip() for x in rows[0]]

        if "已付款金額" not in header and "待付款金額" not in header:
            continue

        paid_idx = header.index("已付款金額") if "已付款金額" in header else None
        unpaid_idx = header.index("待付款金額") if "待付款金額" in header else None

        income_type = detect_income_type(header[0] if header else "")
        source = "儲值金表" if income_type == "儲值金" else "主表"

        for row in rows[1:]:
            if not row:
                continue

            service = normalize_service(row[0] if len(row) > 0 else "")

            if not service:
                continue
            if service == "加總":
                continue
            if service.startswith("LC"):
                continue

            paid = safe_int(row[paid_idx]) if paid_idx is not None and len(row) > paid_idx else 0
            unpaid = safe_int(row[unpaid_idx]) if unpaid_idx is not None and len(row) > unpaid_idx else 0

            results.append({
                "收入類型": income_type,
                "資料來源": source,
                "服務": service,
                "子項目": "",
                "已付款": paid,
                "待付款": unpaid,
            })

    return results


def to_category(service, income) -> Optional[str]:
    if service == "儲值金" and income == "現金收入":
        return "儲值金"
    if service in ["居家清潔", "辦公室清潔", "裝修細清", "搬入清潔", "搬出清潔", "大掃除"]:
        return "清潔"
    if service == "冷氣清潔":
        return "冷氣"
    if service == "洗衣機清潔":
        return "洗衣機"
    if service == "水洗":
        return "水洗"
    if service == "收納":
        return "收納"
    return None


def build_region1_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    work = raw_df.copy()

    work["本月已付款"] = 0
    work["本月待付款"] = 0
    work["下月已付款"] = 0
    work["下月待付款"] = 0

    this_mask = work["月份"] == "本月"
    next_mask = work["月份"] == "下月"

    work.loc[this_mask, "本月已付款"] = work.loc[this_mask, "已付款"]
    work.loc[this_mask, "本月待付款"] = work.loc[this_mask, "待付款"]
    work.loc[next_mask, "下月已付款"] = work.loc[next_mask, "已付款"]
    work.loc[next_mask, "下月待付款"] = work.loc[next_mask, "待付款"]

    region1 = (
        work.groupby(["城市", "收入類型", "資料來源", "服務", "子項目"], as_index=False)[
            ["本月已付款", "本月待付款", "下月已付款", "下月待付款"]
        ]
        .sum()
    )

    region1["城市"] = pd.Categorical(region1["城市"], categories=CITY_ORDER, ordered=True)
    region1["收入類型"] = pd.Categorical(region1["收入類型"], categories=INCOME_ORDER, ordered=True)

    region1 = region1.sort_values(["城市", "收入類型", "服務"]).reset_index(drop=True)
    region1["城市"] = region1["城市"].astype(str)
    region1["收入類型"] = region1["收入類型"].astype(str)

    return region1


def build_region2_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    work = raw_df.copy()
    work["類別"] = work.apply(lambda r: to_category(r["服務"], r["收入類型"]), axis=1)
    work = work[work["類別"].notna()].copy()

    rows = []
    for city in CITY_ORDER:
        for income in INCOME_ORDER:
            for category in CATEGORY_ORDER:
                sub = work[
                    (work["城市"] == city) &
                    (work["收入類型"] == income) &
                    (work["類別"] == category)
                ]

                bm = sub[sub["月份"] == "本月"]
                nm = sub[sub["月份"] == "下月"]

                bm_paid = bm["已付款"].sum()
                bm_unpaid = bm["待付款"].sum()
                nm_paid = nm["已付款"].sum()
                nm_unpaid = nm["待付款"].sum()

                rows.append({
                    "城市": city,
                    "收入類型": income,
                    "類別": category,
                    "本月待付": bm_unpaid,
                    "本月已付": bm_paid,
                    "本月加總": bm_paid + bm_unpaid,
                    "次月待付": nm_unpaid,
                    "次月已付": nm_paid,
                    "次月加總": nm_paid + nm_unpaid,
                })

    return pd.DataFrame(rows)


def build_region3_df(region2_df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for city in CITY_ORDER:
        city_df = region2_df[region2_df["城市"] == city].copy()

        level1 = city_df[["城市", "類別", "收入類型", "本月加總", "次月加總"]].copy()
        level1["加總類型"] = "加總1"
        rows.extend(level1.to_dict("records"))

        mapping_level2 = {
            "清潔現金+儲值金": ["清潔"],
            "家電現金+儲值金": ["冷氣", "洗衣機"],
            "水洗/收納現金+儲值金": ["水洗", "收納"],
        }

        for new_cat, old_cats in mapping_level2.items():
            tmp = city_df[city_df["類別"].isin(old_cats)]
            rows.append({
                "城市": city,
                "類別": new_cat,
                "收入類型": "現金+儲值金",
                "本月加總": tmp["本月加總"].sum(),
                "次月加總": tmp["次月加總"].sum(),
                "加總類型": "加總2",
            })

        mapping_level3 = {
            "清潔+水洗+收納現金+儲值金": ["清潔", "水洗", "收納"],
            "家電現金+儲值金": ["冷氣", "洗衣機"],
        }

        for new_cat, old_cats in mapping_level3.items():
            tmp = city_df[city_df["類別"].isin(old_cats)]
            rows.append({
                "城市": city,
                "類別": new_cat,
                "收入類型": "現金+儲值金",
                "本月加總": tmp["本月加總"].sum(),
                "次月加總": tmp["次月加總"].sum(),
                "加總類型": "加總3",
            })

    region3 = pd.DataFrame(rows)

    type_order = ["加總1", "加總2", "加總3"]

    region3["城市"] = pd.Categorical(region3["城市"], categories=CITY_ORDER, ordered=True)
    region3["加總類型"] = pd.Categorical(region3["加總類型"], categories=type_order, ordered=True)
    region3["類別"] = pd.Categorical(region3["類別"], categories=REGION3_CATEGORY_ORDER, ordered=True)

    region3 = region3.sort_values(["城市", "加總類型", "類別", "收入類型"]).reset_index(drop=True)
    region3["城市"] = region3["城市"].astype(str)
    region3["類別"] = region3["類別"].astype(str)
    region3["收入類型"] = region3["收入類型"].astype(str)
    region3["加總類型"] = region3["加總類型"].astype(str)

    return region3


def build_region4_df(region2_df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for city in CITY_ORDER:
        city_df = region2_df[region2_df["城市"] == city].copy()

        appliance_df = city_df[city_df["類別"].isin(["冷氣", "洗衣機"])]
        bm_appliance = appliance_df["本月加總"].sum()
        nm_appliance = appliance_df["次月加總"].sum()

        cash_stored_df = city_df[
            (city_df["收入類型"] == "現金收入") &
            (city_df["類別"] == "儲值金")
        ]
        bm_cash_stored = cash_stored_df["本月加總"].sum()
        nm_cash_stored = cash_stored_df["次月加總"].sum()

        total_df = city_df[
            ~city_df["類別"].isin(["冷氣", "洗衣機"]) &
            ~(
                (city_df["收入類型"] == "現金收入") &
                (city_df["類別"] == "儲值金")
            )
        ]

        bm_total = total_df["本月加總"].sum()
        nm_total = total_df["次月加總"].sum()

        rows.append({
            "城市": city,
            "本月加總": bm_total,
            "次月加總": nm_total,
            "本月家電加總": bm_appliance,
            "次月家電加總": nm_appliance,
            "儲值金": bm_cash_stored + nm_cash_stored,
        })

    region4 = pd.DataFrame(rows)

    bm_sum = region4["本月加總"].sum()
    nm_sum = region4["次月加總"].sum()

    region4["本月佔比"] = 0 if bm_sum == 0 else region4["本月加總"] / bm_sum
    region4["次月佔比"] = 0 if nm_sum == 0 else region4["次月加總"] / nm_sum

    total_row = pd.DataFrame([{
        "城市": "加總",
        "本月加總": bm_sum,
        "本月佔比": 1,
        "次月加總": nm_sum,
        "次月佔比": 1,
        "本月家電加總": region4["本月家電加總"].sum(),
        "次月家電加總": region4["次月家電加總"].sum(),
        "儲值金": region4["儲值金"].sum(),
    }])

    region4 = pd.concat([region4, total_row], ignore_index=True)

    return region4[[
        "城市",
        "本月加總",
        "本月佔比",
        "次月加總",
        "次月佔比",
        "本月家電加總",
        "次月家電加總",
        "儲值金",
    ]]


def write_four_regions(output_path, df1, df2, df3, df4):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df1.to_excel(writer, index=False, sheet_name="業績報表", startrow=0, startcol=0)
        df2.to_excel(writer, index=False, sheet_name="業績報表", startrow=0, startcol=10)
        df3.to_excel(writer, index=False, sheet_name="業績報表", startrow=0, startcol=20)
        df4.to_excel(writer, index=False, sheet_name="業績報表", startrow=0, startcol=29)

    wb = load_workbook(output_path)
    ws = wb["業績報表"]

    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column in [32, 34]:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "#,##0"

    wb.save(output_path)

def send_region4_email(df4):
    import smtplib
    from email.mime.text import MIMEText
    from datetime import datetime
    import pandas as pd

    sender = "jenny@lemonclean.com.tw"
    password = "bkhe akob wvse ibhm"
    recipient = "jenny@lemonclean.com.tw"

    mail_df = df4.copy()

    # 數字格式（千分位）
    for col in ["本月加總", "次月加總", "本月家電加總", "次月家電加總", "儲值金"]:
        if col in mail_df.columns:
            mail_df[col] = mail_df[col].apply(
                lambda x: f"{int(x):,}" if pd.notna(x) else ""
            )

    # 百分比格式
    if "本月佔比" in mail_df.columns:
        mail_df["本月佔比"] = mail_df["本月佔比"].apply(
            lambda x: f"{x:.2%}" if pd.notna(x) else ""
        )

    if "次月佔比" in mail_df.columns:
        mail_df["次月佔比"] = mail_df["次月佔比"].apply(
            lambda x: f"{x:.2%}" if pd.notna(x) else ""
        )

    html_table = mail_df.to_html(index=False, border=0)

    today_str = datetime.today().strftime("%Y%m%d")
    subject = f"業績報表{today_str}"

    html = f"""
    <html>
      <head>
        <style>
          table {{
            border-collapse: collapse;
            font-family: Arial, sans-serif;
            font-size: 14px;
          }}
          th, td {{
            border: 1px solid #999;
            padding: 6px 10px;
          }}
          th {{
            background-color: #f2f2f2;
            text-align: center;
          }}
          td {{
            text-align: right;
          }}
          td:first-child {{
            text-align: left;
          }}
        </style>
      </head>
      <body>
        <p>您好，以下為業績報表：</p>
        {html_table}
      </body>
    </html>
    """

    msg = MIMEText(html, "html", "utf-8")
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = recipient

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender, password)
        server.sendmail(sender, [recipient], msg.as_string())

    print(f"✅ 已寄出：{recipient}")



def main():
    print("🔥 開始")

    time_tag = sys.argv[1] if len(sys.argv) > 1 else datetime.today().strftime("%H%M")

    (m_start, m_end), (n_start, n_end) = get_ranges()
    merged = {}

    for city in CITY_ORDER:
        print(f"\n===== {city} =====")

        session = requests.Session()
        acc = ACCOUNTS[city]

        try:
            login(session, acc["email"], acc["password"])

            for label, (s, e) in {
                "本月": (m_start, m_end),
                "下月": (n_start, n_end),
            }.items():
                for status in [1, 0]:
                    for kw in get_keywords(city):
                        url = build_url(s, e, status, kw)
                        res = session.get(url, headers=HEADERS, allow_redirects=True)
                        res.raise_for_status()

                        rows = parse_html(res.text)

                        for row in rows:
                            key = (
                                city,
                                label,
                                row["收入類型"],
                                row["資料來源"],
                                row["服務"],
                                row["子項目"],
                            )

                            if key not in merged:
                                merged[key] = {
                                    "城市": city,
                                    "月份": label,
                                    "收入類型": row["收入類型"],
                                    "資料來源": row["資料來源"],
                                    "服務": row["服務"],
                                    "子項目": row["子項目"],
                                    "已付款": 0,
                                    "待付款": 0,
                                }

                            merged[key]["已付款"] += row["已付款"]
                            merged[key]["待付款"] += row["待付款"]

        except Exception as e:
            print(f"❌ {city} 失敗：{e}")

    raw_df = pd.DataFrame(merged.values())
    if raw_df.empty:
        raise RuntimeError("沒有任何資料可輸出")

    df1 = build_region1_df(raw_df)
    df2 = build_region2_df(raw_df)
    df3 = build_region3_df(df2)
    df4 = build_region4_df(df2)

    save_dir = f"{PATH_JENNY}/業績報表"
    os.makedirs(save_dir, exist_ok=True)

    output = os.path.join(
        save_dir,
        f"業績報表_{datetime.today().strftime('%Y%m%d')}_{time_tag}.xlsx"
    )

    write_four_regions(output, df1, df2, df3, df4)
    print(f"✅ 已輸出：{output}")

    send_region4_email(df4)

if __name__ == "__main__":
    main()
